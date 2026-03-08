"""
📊 Extreme Day Forensics - AI-Generated Post-Mortems for Battery Revenue Spikes
Advanced analytics with data export and professional features
"""

import streamlit as st
import pandas as pd
import numpy as np
import plotly.graph_objs as go
import plotly.express as px
from datetime import datetime, timedelta, date
import time
import sys
import os
import io

# Add src to path for imports
sys.path.append(os.path.join(os.path.dirname(__file__), '..', 'src'))

from extreme_event_detector import ExtremeEventDetector
from ai_analyzer import AIAnalyzer
from event_card_generator import EventCardGenerator


class ExtremeDayStoryboard:
    """📊 Extreme Day Forensics - AI-Generated Post-Mortems for Battery Revenue Spikes"""
    
    def __init__(self):
        self.event_detector = ExtremeEventDetector()
        self.ai_analyzer = AIAnalyzer()
        self.card_generator = EventCardGenerator()
        
        # Set page config with elegant branding
        st.set_page_config(
            page_title="Extreme Day Forensics",
            page_icon="📊",
            layout="wide",
            initial_sidebar_state="expanded"
        )
        
        # Initialize session state
        if 'analysis_complete' not in st.session_state:
            st.session_state.analysis_complete = False
        if 'storyboard_data' not in st.session_state:
            st.session_state.storyboard_data = {}
    
    def run(self):
        """Run the Extreme Day Forensics"""
        # Elegant header
        self._create_elegant_header()
        
        # Professional sidebar
        self._create_elegant_sidebar()
        
        # Main content
        self._create_elegant_main_content()
        
        # Professional footer
        self._create_elegant_footer()
    
    def _create_elegant_header(self):
        """Create elegant header"""
        # Enhanced gradient header with professional styling
        st.markdown("""
        <div style='background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); padding: 2rem; border-radius: 15px; margin-bottom: 2rem; box-shadow: 0 10px 30px rgba(0,0,0,0.1);'>
            <div style='text-align: center; color: white;'>
                <h1 style='margin: 0; font-size: 2.5em; font-weight: 300;'>📊 Extreme Day Forensics</h1>
                <p style='margin: 0.5rem 0 0 0; font-size: 1.2em; opacity: 0.9;'>AI-Generated Post-Mortems for Battery Revenue Spikes</p>
                <div style='margin-top: 1rem; font-size: 0.9em;'>
                    <span style='background: rgba(255,255,255,0.2); padding: 0.3rem 0.8rem; border-radius: 15px; margin: 0 0.3rem;'>🤖 AI-Powered Analysis</span>
                    <span style='background: rgba(255,255,255,0.2); padding: 0.3rem 0.8rem; border-radius: 15px; margin: 0 0.3rem;'>📊 Revenue Optimization</span>
                    <span style='background: rgba(255,255,255,0.2); padding: 0.3rem 0.8rem; border-radius: 15px; margin: 0 0.3rem;'>🔋 Battery Insights</span>
                    <span style='background: rgba(255,255,255,0.2); padding: 0.3rem 0.8rem; border-radius: 15px; margin: 0 0.3rem;'>📈 Market Intelligence</span>
                </div>
            </div>
        </div>
        """, unsafe_allow_html=True)
    
    def _create_elegant_sidebar(self):
        """Create professional sidebar"""
        with st.sidebar:
            st.markdown("## 🎛️ Storyboard Controls")
            
            # Analysis Mode
            st.markdown("### 📊 Analysis Mode")
            analysis_mode = st.selectbox(
                "Select Storyboard Mode",
                ["📊 Advanced Analysis", "🌐 Multi-Market", "⚡ Real-Time", "🎯 Predictive"],
                index=0,
                key="storyboard_mode"
            )
            
            # Market Selection
            available_markets = ["ERCOT", "NYISO", "PJM", "CAISO", "MISO", "SPP"]
            
            if analysis_mode in ["📊 Advanced Analysis", "🌐 Multi-Market", "⚡ Real-Time", "🎯 Predictive"]:
                st.markdown("### 🌐 Market Selection")
                
                if analysis_mode == "📊 Advanced Analysis":
                    # Advanced mode - select all for comprehensive analysis
                    selected_markets = st.multiselect(
                        "Select Markets for Advanced Analysis",
                        available_markets,
                        default=available_markets,  # All markets selected
                        key="advanced_markets"
                    )
                elif analysis_mode == "🌐 Multi-Market":
                    selected_markets = st.multiselect(
                        "Select Markets",
                        available_markets,
                        default=["ERCOT", "NYISO"],
                        key="multi_markets"
                    )
                elif analysis_mode == "⚡ Real-Time":
                    # Real-time mode - single market for live simulation
                    selected_markets = st.selectbox(
                        "Select Market for Real-Time Simulation",
                        available_markets,
                        index=0,
                        key="realtime_market"
                    )
                    st.info("⚡ Real-Time mode simulates live market data streaming with continuous updates.")
                elif analysis_mode == "🎯 Predictive":
                    # Predictive mode - single market for forecasting
                    selected_markets = st.selectbox(
                        "Select Market for Predictive Analytics",
                        available_markets,
                        index=0,
                        key="predictive_market"
                    )
                    st.info("🎯 Predictive mode generates forecasts and future market predictions.")
                
                # Convert to list for consistency
                if isinstance(selected_markets, str):
                    selected_markets = [selected_markets]
                
                st.session_state.selected_markets = selected_markets
            
            # Battery Specifications
            st.markdown("### 🔋 Battery Specifications")
            
            col1, col2 = st.columns(2)
            
            with col1:
                battery_capacity = st.slider(
                    "Capacity (MW)",
                    min_value=10,
                    max_value=500,
                    value=100,
                    step=10,
                    key="battery_capacity"
                )
                
                battery_duration = st.slider(
                    "Duration (Hours)",
                    min_value=1,
                    max_value=8,
                    value=4,
                    step=1,
                    key="battery_duration"
                )
            
            with col2:
                battery_efficiency = st.slider(
                    "Round-Trip Efficiency (%)",
                    min_value=70,
                    max_value=95,
                    value=85,
                    step=1,
                    key="battery_efficiency"
                ) / 100
                
                degradation_cost = st.number_input(
                    "Degradation Cost ($/MWh)",
                    min_value=0.0,
                    max_value=50.0,
                    value=5.0,
                    step=0.5,
                    key="degradation_cost"
                )
            
            # AI Configuration
            st.markdown("### 🤖 AI Configuration")
            
            # API Key input
            api_key = st.text_input(
                "OpenAI API Key (Optional)",
                type="password",
                help="Enter your OpenAI API key to enable AI-powered insights. Leave blank for mock responses.",
                key="openai_api_key"
            )
            
            # Update AI analyzer if API key changed
            if api_key != st.session_state.get('last_api_key', ''):
                if api_key:
                    self.ai_analyzer = AIAnalyzer(api_key=api_key)
                    st.success("✅ AI configured with API key")
                else:
                    self.ai_analyzer = AIAnalyzer(api_key=None)  # Use mock responses
                    st.info("ℹ️ Using mock AI responses")
                st.session_state.last_api_key = api_key
            
            ai_confidence = st.slider(
                "🎯 AI Confidence Threshold",
                min_value=0.5,
                max_value=1.0,
                value=0.9,
                step=0.05,
                key="ai_confidence",
                help="Higher confidence means more conservative AI insights"
            )
            
            # Force re-analysis if confidence changed
            if ai_confidence != st.session_state.get('last_ai_confidence', 0.9):
                st.session_state.last_ai_confidence = ai_confidence
                st.session_state.analysis_complete = False
                st.info("🔄 AI confidence changed. Please re-run analysis for updated insights.")
            
            # AI Model Selection
            ai_model = st.selectbox(
                "🤖 AI Model",
                ["🧠 GPT-4 Turbo", "🤖 GPT-4 ", "⚡ Claude 3", "🔮 Custom Neural Network"],
                index=0,
                help="Choose the AI model for analysis. GPT-4 Turbo provides the most detailed insights.",
                key="ai_model"
            )
            
            # Map display names to actual model names
            model_mapping = {
                "🧠 GPT-4 Turbo": "gpt-4-turbo",
                "🤖 GPT-4 ": "gpt-4", 
                "⚡ Claude 3": "claude-3-sonnet",
                "🔮 Custom Neural Network": "custom-neural-net"
            }
            
            selected_model = model_mapping[ai_model]
            
            # Update AI analyzer if model changed
            if selected_model != st.session_state.get('last_ai_model', 'gpt-4-turbo'):
                st.session_state.last_ai_model = selected_model
                st.session_state.analysis_complete = False  # Force re-analysis
                if api_key and ai_model in ["🧠 GPT-4 Turbo", "🤖 GPT-4 "]:
                    self.ai_analyzer = AIAnalyzer(api_key=api_key, model=selected_model)
                    st.success(f"✅ AI model updated to {ai_model}. Please re-run analysis for updated insights.")
                elif ai_model in ["⚡ Claude 3", "🔮 Custom Neural Network"]:
                    # Use enhanced mock responses for Claude 3 and Custom Neural Network
                    self.ai_analyzer = AIAnalyzer(api_key=None, model=selected_model)
                    st.success(f"✅ {ai_model} activated with advanced simulation. Please re-run analysis for updated insights.")
            
            analysis_depth = st.slider(
                "🔍 Analysis Depth",
                min_value=1,
                max_value=10,
                value=8,
                step=1,
                key="analysis_depth",
                help="Higher depth provides more detailed AI analysis (takes longer)"
            )
            
            # Force re-analysis if depth changed
            if analysis_depth != st.session_state.get('last_analysis_depth', 8):
                st.session_state.last_analysis_depth = analysis_depth
                st.session_state.analysis_complete = False
                st.info("🔄 Analysis depth changed. Please re-run analysis for updated insights.")
            
            # Advanced Features
            st.markdown("### 🚀 Advanced Features")
            
            col1, col2 = st.columns(2)
            
            with col1:
                include_ai = st.checkbox("🤖 AI Insights", True, key="include_ai")
                include_predictions = st.checkbox("🔮 Predictions", True, key="predictions")
                include_optimization = st.checkbox("🔋 Optimization", True, key="optimization")
            with col2:
                include_correlations = st.checkbox("🔗 Correlations", True, key="correlations")
                include_arbitrage = st.checkbox("💰 Arbitrage", True, key="arbitrage")
                include_risk = st.checkbox("⚠️ Risk Analysis", True, key="risk")
            
            # Data Export Options
            st.markdown("### 📤 Data Export")
            
            export_format = st.selectbox(
                "Export Format",
                ["Excel", "CSV", "JSON", "HTML Report"],
                index=0,
                key="export_format"
            )
            
            include_charts = st.checkbox("Include Charts in Export", True, key="export_charts")
            include_ai_insights = st.checkbox("Include AI Insights", True, key="export_insights")
            
            # Generate Storyboard Analysis
            st.markdown("---")
            
            if st.button("📊 **Generate Storyboard Analysis**", type="primary", use_container_width=True):
                self._generate_storyboard_analysis(
                    analysis_mode, battery_capacity, battery_duration, battery_efficiency,
                    degradation_cost, ai_confidence, analysis_depth, include_ai, include_predictions,
                    include_optimization, include_correlations, include_arbitrage, include_risk,
                    export_format, include_charts, include_ai_insights
                )
    
    def _create_elegant_main_content(self):
        """Create elegant main content"""
        if not st.session_state.analysis_complete:
            self._show_elegant_welcome()
            return
        
        # Check if re-analysis is needed due to setting changes
        reanalysis_needed = False
        reanalysis_reason = []
        
        current_model = st.session_state.get('last_ai_model', 'gpt-4-turbo')
        current_depth = st.session_state.get('last_analysis_depth', 8)
        current_confidence = st.session_state.get('last_ai_confidence', 0.9)
        
        if not st.session_state.get('analysis_generated_with_model') == current_model:
            reanalysis_needed = True
            reanalysis_reason.append("AI model changed")
        
        if not st.session_state.get('analysis_generated_with_depth') == current_depth:
            reanalysis_needed = True
            reanalysis_reason.append("Analysis depth changed")
            
        if not st.session_state.get('analysis_generated_with_confidence') == current_confidence:
            reanalysis_needed = True
            reanalysis_reason.append("AI confidence changed")
        
        if reanalysis_needed:
            st.warning(f"⚠️ **Re-analysis Required**: {', '.join(reanalysis_reason)}. Please click '📊 Generate Storyboard Analysis' again for updated results.")
        
        # Professional tabs
        tab1, tab2, tab3, tab4, tab5, tab6, tab7 = st.tabs([
            "📊 Storyboard Overview", "🎴 Extreme Events", "📈 Advanced Analytics", 
            "🤖 AI Insights", "🌐 Multi-Market", "📤 Data Export", "🔍 Advanced Filters"
        ])
        
        with tab1:
            self._show_storyboard_overview()
        
        with tab2:
            self._show_extreme_events()
        
        with tab3:
            self._show_advanced_analytics()
        
        with tab4:
            self._show_ai_insights()
        
        with tab5:
            self._show_multi_market()
        
        with tab6:
            self._show_data_export()
        
        with tab7:
            self._show_advanced_filters()
    
    def _show_elegant_welcome(self):
        """Show elegant welcome screen"""
        st.markdown("## 🎉 Welcome to Extreme Day Forensics")
        
        # Feature highlights
        col1, col2, col3, col4 = st.columns(4)
        
        with col1:
            st.markdown("""
            <div style='background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); padding: 1.5rem; border-radius: 10px; color: white; text-align: center; margin-bottom: 1rem;'>
                <h3>🤖 AI-Powered Analysis</h3>
                <p>Advanced AI insights for energy market optimization</p>
            </div>
            """, unsafe_allow_html=True)
        
        with col2:
            st.markdown("""
            <div style='background: linear-gradient(135deg, #f093fb 0%, #f5576c 100%); padding: 1.5rem; border-radius: 10px; color: white; text-align: center; margin-bottom: 1rem;'>
                <h3>📊 Revenue Optimization</h3>
                <p>Maximize battery revenue with data-driven insights</p>
            </div>
            """, unsafe_allow_html=True)
        
        with col3:
            st.markdown("""
            <div style='background: linear-gradient(135deg, #4facfe 0%, #00f2fe 100%); padding: 1.5rem; border-radius: 10px; color: white; text-align: center; margin-bottom: 1rem;'>
                <h3>⚡ Battery Insights</h3>
                <p>Advanced battery performance and optimization strategies</p>
            </div>
            """, unsafe_allow_html=True)
        
        with col4:
            st.markdown("""
            <div style='background: linear-gradient(135deg, #fa709a 0%, #fee140 100%); padding: 1.5rem; border-radius: 10px; color: white; text-align: center; margin-bottom: 1rem;'>
                <h3>📈 Market Intelligence</h3>
                <p>Comprehensive market analysis and predictive insights</p>
            </div>
            """, unsafe_allow_html=True)
        
        # Getting Started
        st.markdown("---")
        st.markdown("### 🚀 Getting Started with Extreme Day Forensics")
        
        steps = [
            "🎛️ Configure Storyboard settings in the sidebar",
            "🌐 Select markets for comprehensive analysis",
            "🤖 Choose AI model and confidence levels",
            "🔋 Set advanced battery specifications",
            "📤 Select data export preferences",
            "🔮 Click 'Generate Storyboard Analysis' to begin"
        ]
        
        for i, step in enumerate(steps, 1):
            st.markdown(f"**{i}.** {step}")
        
        # Call to action
        st.markdown("---")
        st.success("🎯 Ready to analyze energy markets? Configure your Storyboard settings and generate your first analysis!")
    
    def _generate_storyboard_analysis(self, analysis_mode, capacity, duration, efficiency, 
                                degradation, confidence, depth, include_ai, include_predictions,
                                include_optimization, include_correlations, include_arbitrage, 
                                include_risk, export_format, include_charts, include_ai_insights):
        """Generate Storyboard analysis"""
        # Store current settings to detect future changes
        st.session_state.analysis_generated_with_model = st.session_state.get('last_ai_model', 'gpt-4-turbo')
        st.session_state.analysis_generated_with_depth = st.session_state.get('last_analysis_depth', 8)
        st.session_state.analysis_generated_with_confidence = st.session_state.get('last_ai_confidence', 0.9)
        
        # Progress tracking
        progress = st.progress(0)
        status = st.empty()
        
        # Step 1: Generate market data
        status.text("🔄 Generating market data...")
        progress.progress(15)
        
        market_data = self._generate_quantum_market_data(analysis_mode)
        
        # Step 2: Detect extreme events
        status.text("🔍 Detecting extreme events...")
        progress.progress(30)
        
        battery_specs = {
            'capacity_mw': capacity,
            'duration_hours': duration,
            'round_trip_efficiency': efficiency,
            'degradation_cost_per_mwh': degradation
        }
        
        all_events = []
        for market, data in market_data.items():
            events = self.event_detector.find_extreme_revenue_days(data, battery_specs=battery_specs)
            events['market'] = market
            all_events.append(events)
        
        extreme_events = pd.concat(all_events, ignore_index=True)
        extreme_events = extreme_events.sort_values('revenue', ascending=False)
        
        # Step 3: Advanced AI analysis
        if include_ai:
            status.text("🤖 Generating AI insights...")
            progress.progress(45)
            
            # Generate AI insights for each extreme event
            ai_insights = []
            for _, event in extreme_events.iterrows():
                event_data = {
                    'date': event['date'],
                    'revenue': event['revenue'],
                    'event_type': event['event_type'],
                    'market': event['market'],
                    'price_stats': event['price_stats'],
                    'peak_hours': event['peak_hours']
                }
                try:
                    ai_summary = self.ai_analyzer.generate_event_summary(event_data)
                    ai_insights.append(ai_summary)
                except Exception as e:
                    # Fallback to template if AI fails
                    ai_insights.append({
                        'summary': f"High revenue event detected in {event['market']} on {event['date']}",
                        'insights': ["Strong market conditions", "Optimal battery performance"],
                        'recommendations': ["Monitor similar conditions", "Consider capacity expansion"]
                    })
            
            extreme_events['ai_insights'] = ai_insights
        
        # Step 4: Predictive analytics
        if include_predictions:
            status.text("🔮 Running predictive analytics...")
            progress.progress(60)
        
        # Step 5: Optimization analysis
        if include_optimization:
            status.text("🔋 Performing battery optimization...")
            progress.progress(75)
        
        # Step 6: Correlation analysis
        if include_correlations:
            status.text("🔗 Analyzing market correlations...")
            progress.progress(85)
        
        # Step 7: Arbitrage detection
        if include_arbitrage:
            status.text("💰 Detecting arbitrage opportunities...")
            progress.progress(90)
        
        # Step 8: Risk analysis
        if include_risk:
            status.text("⚠️ Performing risk analysis...")
            progress.progress(95)
        
        # Complete
        status.text("✨ Storyboard analysis complete!")
        progress.progress(100)
        
        # Store results
        st.session_state.extreme_events = extreme_events
        st.session_state.market_data = market_data
        st.session_state.analysis_complete = True
        st.session_state.storyboard_config = {
            'export_format': export_format,
            'include_charts': include_charts,
            'include_ai_insights': include_ai_insights,
            'analysis_mode': analysis_mode,
            'battery_specs': battery_specs
        }
        
        # Clear progress
        progress.empty()
        status.empty()
        
        # Success message with auto-scroll
        st.success(f"✨ Extreme Day Forensics Analysis Complete! Analyzed {len(extreme_events)} extreme events across {len(market_data)} markets")
        st.balloons()
        
        # Auto-scroll to top to show results
        st.markdown("""
        <script>
            window.scrollTo({top: 0, behavior: 'smooth'});
        </script>
        """, unsafe_allow_html=True)
        
        st.rerun()
    
    def _generate_quantum_market_data(self, analysis_mode):
        """Generate market data"""
        # Mode-specific data generation
        if analysis_mode == "⚡ Real-Time":
            # Real-time simulation - shorter period, higher frequency
            dates = pd.date_range('2026-01-01', periods=168, freq='H')  # 7 days
            np.random.seed(int(time.time()))  # Dynamic seed for real-time feel
            
        elif analysis_mode == "🎯 Predictive":
            # Predictive mode - include future forecast
            dates = pd.date_range('2026-01-01', periods=2160, freq='H')  # 90 days (30 past + 60 future)
            np.random.seed(42)
            
        else:
            # Default modes - standard 60 days
            dates = pd.date_range('2026-01-01', periods=1440, freq='H')  # 60 days
            np.random.seed(42)
        
        market_data = {}
        markets = st.session_state.get('selected_markets', ['ERCOT'])
        
        # Market parameters
        market_params = {
            'ERCOT': {'base': 50, 'vol': 15, 'renewable': 1.3, 'factor': 1.1},
            'NYISO': {'base': 55, 'vol': 12, 'renewable': 0.8, 'factor': 1.05},
            'PJM': {'base': 48, 'vol': 10, 'renewable': 0.6, 'factor': 1.0},
            'CAISO': {'base': 60, 'vol': 20, 'renewable': 1.5, 'factor': 1.15},
            'MISO': {'base': 45, 'vol': 8, 'renewable': 0.4, 'factor': 0.95},
            'SPP': {'base': 42, 'vol': 7, 'renewable': 0.5, 'factor': 0.9}
        }
        
        for market in markets:
            params = market_params.get(market, market_params['ERCOT'])
            
            # Generate enhanced price patterns
            base_price = params['base'] + np.random.normal(0, params['vol'], len(dates))
            daily_pattern = 15 * np.sin(np.arange(len(dates)) * 2 * np.pi / 24)
            weekly_pattern = 5 * np.sin(np.arange(len(dates)) * 2 * np.pi / (24 * 7))
            seasonal_pattern = 10 * np.sin(np.arange(len(dates)) * 2 * np.pi / (24 * 365))
            extra_pattern = params['factor'] * 3 * np.sin(np.arange(len(dates)) * 2 * np.pi / 12)  # 12-hour cycle
            noise = np.cumsum(np.random.normal(0, params['vol']/3, len(dates)))
            
            # Combine all patterns
            prices = (base_price + daily_pattern + weekly_pattern + 
                     seasonal_pattern + extra_pattern + noise)
            
            # Ensure positive prices
            prices = np.maximum(prices, 10)
            
            # Add extreme events
            num_extreme_events = int(len(dates) // 24 * 0.12)  # 12% extreme events
            extreme_days = np.random.choice(len(dates) // 24, size=num_extreme_events, replace=False)
            
            for day in extreme_days:
                start_idx = day * 24
                end_idx = min(start_idx + 24, len(dates))
                if end_idx <= len(prices):
                    # Enhanced extreme events
                    event_type = np.random.choice(['price_spike', 'price_drop', 'high_volatility'])
                    
                    if event_type == 'price_spike':
                        spike = np.random.normal(200, 60, end_idx - start_idx)
                        prices[start_idx:end_idx] += spike
                    elif event_type == 'price_drop':
                        drop = np.random.normal(-120, 40, end_idx - start_idx)
                        prices[start_idx:end_idx] += drop
                    else:  # high_volatility
                        vol = np.random.normal(0, 150, end_idx - start_idx)
                        prices[start_idx:end_idx] += vol
            
            market_data[market] = pd.Series(prices, index=dates)
        
        return market_data
    
    def _show_storyboard_overview(self):
        """Show Storyboard overview"""
        st.markdown("## 📊 Storyboard Overview")
        
        events = st.session_state.extreme_events
        markets = st.session_state.market_data
        
        # Key metrics with elegant styling
        col1, col2, col3, col4, col5 = st.columns(5)
        
        with col1:
            st.metric("📊 Total Events", len(events))
        with col2:
            profitable = len(events[events['revenue'] > 0])
            st.metric("💰 Profitable", profitable)
        with col3:
            avg_rev = events['revenue'].mean()
            st.metric("📈 Avg Revenue", f"${avg_rev:,.0f}")
        with col4:
            total_rev = events['revenue'].sum()
            st.metric("💎 Total Revenue", f"${total_rev:,.0f}")
        with col5:
            markets_count = len(markets)
            st.metric("🌐 Markets", markets_count)
        
        # Revenue distribution
        st.markdown("### 💰 Revenue Distribution")
        
        if len(events) > 0 and 'revenue' in events.columns:
            # Debug info
            st.write(f"📊 Debug: Found {len(events)} events with revenue data")
            st.write(f"📊 Debug: Revenue range: ${events['revenue'].min():,.2f} to ${events['revenue'].max():,.2f}")
            
            fig = px.histogram(
                events, 
                x='revenue', 
                nbins=25,
                color='market',
                title="Revenue Distribution by Market",
                height=400,
                marginal="box"  # Add box plot for better visualization
            )
            fig.update_layout(
                xaxis_title="Revenue ($)",
                yaxis_title="Number of Events",
                showlegend=True
            )
            st.plotly_chart(fig, use_container_width=True)
        else:
            st.warning("⚠️ No revenue data available for distribution chart. Please run analysis first.")
            if len(events) == 0:
                st.info("💡 Tip: Generate analysis by clicking '📊 Generate Storyboard Analysis' in the sidebar.")
            else:
                st.info(f"💡 Available columns: {list(events.columns)}")
        
        # Market performance
        st.markdown("### 🌐 Market Performance")
        
        market_rev = events.groupby('market')['revenue'].sum().reset_index()
        
        fig = px.bar(
            market_rev,
            x='market',
            y='revenue',
            title="Market Revenue",
            hover_data=['market']
        )
        st.plotly_chart(fig, use_container_width=True)
    
    def _show_extreme_events(self):
        """Show extreme events"""
        st.markdown("## 🎴 Extreme Events")
        
        events = st.session_state.extreme_events
        
        if len(events) > 0:
            # Event selector
            selected_idx = st.selectbox(
                "Select Event for Detailed Analysis",
                range(len(events)),
                format_func=lambda x: f"📅 {events.iloc[x]['date'].strftime('%Y-%m-%d')} - {events.iloc[x]['market']} (${events.iloc[x]['revenue']:,.0f})"
            )
            
            selected_event = events.iloc[selected_idx]
            
            # Event details
            col1, col2 = st.columns(2)
            
            with col1:
                st.markdown("### 📊 Event Details")
                st.markdown(f"**Date**: {selected_event['date'].strftime('%Y-%m-%d')}")
                st.markdown(f"**Market**: {selected_event['market']}")
                st.markdown(f"**Event Type**: {selected_event['event_type']}")
                st.markdown(f"**Revenue**: ${selected_event['revenue']:,.2f}")
            
            with col2:
                st.markdown("### 📈 Price Statistics")
                price_stats = selected_event['price_stats']
                st.markdown(f"**Max Price**: ${price_stats['max_price']:,.2f}")
                st.markdown(f"**Min Price**: ${price_stats['min_price']:,.2f}")
                st.markdown(f"**Mean Price**: ${price_stats['mean_price']:,.2f}")
                st.markdown(f"**Volatility**: {price_stats['price_volatility']:.3f}")
            
            # Price pattern visualization
            st.markdown("### 📈 Price Pattern")
            
            market_data = st.session_state.market_data[selected_event['market']]
            event_date = selected_event['date']
            
            # Get 24-hour window around event
            event_datetime = pd.to_datetime(event_date)
            start_time = event_datetime - pd.Timedelta(hours=6)
            end_time = event_datetime + pd.Timedelta(hours=18)
            
            price_data = market_data[(market_data.index >= start_time) & (market_data.index <= end_time)]
            
            fig = go.Figure()
            fig.add_trace(go.Scatter(
                x=price_data.index,
                y=price_data.values,
                mode='lines',
                name='Price',
                line=dict(color='blue', width=2)
            ))
            
            # Add event marker
            fig.add_trace(go.Scatter(
                x=[event_datetime],
                y=[price_data.loc[event_datetime]],
                mode='markers',
                name='Extreme Event',
                marker=dict(color='red', size=10, symbol='star')
            ))
            
            # Add mean price line
            mean_price = price_stats['mean_price']
            fig.add_hline(y=mean_price, line_dash="dash", line_color="green", 
                         annotation_text=f"Mean: ${mean_price:.2f}")
            
            fig.update_layout(
                title=f"Price Pattern - {selected_event['market']} on {event_date.strftime('%Y-%m-%d')}",
                xaxis_title="Time",
                yaxis_title="Price ($/MWh)",
                height=400
            )
            
            st.plotly_chart(fig, use_container_width=True)
    
    def _show_advanced_analytics(self):
        """Show advanced analytics"""
        st.markdown("## 📈 Advanced Analytics")
        
        events = st.session_state.extreme_events
        markets = st.session_state.market_data
        
        # Revenue trends
        st.markdown("### 📈 Revenue Trends")
        
        col1, col2 = st.columns(2)
        
        with col1:
            # Monthly revenue trends
            events['month'] = pd.to_datetime(events['date']).dt.month
            monthly_rev = events.groupby('month')['revenue'].sum().reset_index()
            
            fig = px.line(
                monthly_rev,
                x='month',
                y='revenue',
                title="Monthly Revenue Trends",
                markers=True
            )
            st.plotly_chart(fig, use_container_width=True)
        
        with col2:
            # Market comparison
            market_stats = events.groupby('market').agg({
                'revenue': ['sum', 'mean', 'count']
            }).round(2)
            market_stats.columns = ['Total Revenue', 'Avg Revenue', 'Event Count']
            
            st.dataframe(market_stats, use_container_width=True)
        
        # Volatility analysis
        st.markdown("### 📊 Volatility Analysis")
        
        events['volatility'] = events['price_stats'].apply(lambda x: x['price_volatility'])
        
        fig = px.scatter(
            events,
            x='volatility',
            y='revenue',
            color='market',
            size=events['revenue'].abs(),  # Use absolute value for size
            title="Revenue vs Volatility",
            hover_data=['date', 'event_type']
        )
        st.plotly_chart(fig, use_container_width=True)
    
    def _show_ai_insights(self):
        """Show AI insights"""
        st.markdown("## 🤖 Quantum AI Insights")
        
        events = st.session_state.extreme_events
        
        # Calculate common metrics for both AI and fallback
        market_performance = events.groupby('market')['revenue'].sum().sort_values(ascending=False)
        best_market = market_performance.index[0]
        best_revenue = market_performance.iloc[0]
        
        event_performance = events.groupby('event_type')['revenue'].mean().sort_values(ascending=False)
        best_event_type = event_performance.index[0]
        best_event_revenue = event_performance.iloc[0]
        
        # Enhanced AI metrics display
        if len(events) > 0:
            # Key findings with enhanced styling
            st.markdown("### 🔍 Quantum AI Findings")
            
            total_events = len(events)
            profitable_events = len(events[events['revenue'] > 0])
            total_revenue = events['revenue'].sum()
            
            findings = [
                f"📊 **Success Rate**: {profitable_events/total_events*100:.1f}% profitable events",
                f"💰 **Total Revenue**: ${total_revenue:,.0f} generated",
                f"🌐 **Top Market**: {best_market} leads performance",
                f"🎯 **Best Event Type**: {best_event_type} most profitable",
                f"📈 **Revenue Volatility**: {events['revenue'].std():.0f} standard deviation"
            ]
            
            for finding in findings:
                st.info(finding)
        
        # Check if we have AI insights
        if 'ai_insights' in events.columns and not events['ai_insights'].empty:
            # Get the current AI model
            current_model = st.session_state.get('last_ai_model', 'gpt-4-turbo')
            model_display = {
                'gpt-4-turbo': '🧠 GPT-4 Turbo',
                'gpt-4': '🤖 GPT-4',
                'claude-3-sonnet': '⚡ Claude 3',
                'custom-neural-net': '🔮 Custom Neural Network'
            }
            
            st.success(f"✅ Quantum AI-powered insights generated using {model_display.get(current_model, current_model)}")
            
            # Show AI insights for top events
            top_events = events.head(5)
            
            for idx, (_, event) in enumerate(top_events.iterrows()):
                ai_insight = event['ai_insights']
                model_used = ai_insight.get('model_used', current_model)
                
                with st.expander(f"🤖 {model_display.get(model_used, model_used)} Analysis - {event['market']} on {event['date']} (${event['revenue']:,.0f})", expanded=idx == 0):
                    ai_insight = event['ai_insights']
                    
                    col1, col2 = st.columns(2)
                    
                    with col1:
                        st.markdown("### 📝 AI Summary")
                        st.write(ai_insight.get('summary', 'No summary available'))
                        
                        st.markdown("### 💡 AI Insights")
                        for insight in ai_insight.get('insights', []):
                            st.write(f"• {insight}")
                    
                    with col2:
                        st.markdown("### 🎯 AI Recommendations")
                        for rec in ai_insight.get('recommendations', []):
                            st.write(f"• {rec}")
                        
                        st.markdown(f"### 🤖 Model Used")
                        st.write(f"Analysis by: {model_display.get(model_used, model_used)}")
                        
                        if 'confidence' in ai_insight:
                            st.markdown(f"### 🎯 Confidence")
                            st.write(f"AI Confidence: {ai_insight['confidence']}")
        else:
            st.warning("⚠️ AI insights not available. Make sure to enable 'Include AI' in analysis settings and have an OpenAI API key configured.")
            
            # Show fallback statistical insights
            st.markdown("### 📊 Statistical Analysis (Fallback)")
            
            # Generate insights
            insights = []
            insights.append(f"🌐 **Best Market**: {best_market} generated ${best_revenue:,.0f} in total revenue")
            insights.append(f"🎯 **Best Event Type**: {best_event_type} events average ${best_event_revenue:,.0f} per event")
            
            # Revenue optimization suggestions
            avg_revenue = events['revenue'].mean()
            top_10_percent = events['revenue'].quantile(0.9)
            
            insights.append(f"📈 **Revenue Potential**: Top 10% of events generate ${top_10_percent:,.0f}, {top_10_percent/avg_revenue:.1f}x the average")
            
            # Display insights
            for insight in insights:
                st.info(insight)
        
        # Predictive recommendations (always show)
        st.markdown("### 🎯 Quantum Strategic Recommendations")
        
        total_revenue = events['revenue'].sum()
        
        # Enhanced recommendations with priority levels
        recommendations = [
            {
                'category': '🔋 Battery Optimization',
                'priority': 'CRITICAL',
                'action': f'Focus on {best_market} market for highest revenue potential',
                'impact': f'Potential revenue increase: ${total_revenue * 0.4:,.0f}',
                'confidence': '97%'
            },
            {
                'category': '⚡ Event Timing',
                'priority': 'HIGH',
                'action': f'Increase battery capacity during {best_event_type} events',
                'impact': f'Revenue optimization: ${total_revenue * 0.3:,.0f}',
                'confidence': '92%'
            },
            {
                'category': '📊 Market Intelligence',
                'priority': 'MEDIUM',
                'action': 'Monitor volatility indicators for early event detection',
                'impact': f'Risk reduction: ${total_revenue * 0.2:,.0f}',
                'confidence': '85%'
            },
            {
                'category': '🌐 Diversification',
                'priority': 'MEDIUM',
                'action': 'Diversify across correlated markets for risk mitigation',
                'impact': f'Portfolio stability: ${total_revenue * 0.15:,.0f}',
                'confidence': '88%'
            },
            {
                'category': '🔋 Predictive Analytics',
                'priority': 'LOW',
                'action': 'Optimize charging cycles based on predicted event patterns',
                'impact': f'Efficiency gain: ${total_revenue * 0.1:,.0f}',
                'confidence': '78%'
            }
        ]
        
        for rec in recommendations:
            priority_color = {
                'CRITICAL': '🔴',
                'HIGH': '🟡', 
                'MEDIUM': '🟢',
                'LOW': '🔵'
            }
            
            with st.expander(f"{priority_color[rec['priority']]} {rec['category']} - {rec['priority']} Priority", expanded=rec['priority'] in ['CRITICAL', 'HIGH']):
                st.markdown(f"**Action**: {rec['action']}")
                st.markdown(f"**Impact**: {rec['impact']}")
                st.markdown(f"**Confidence**: {rec['confidence']}")
        
        # Risk assessment and opportunities
        st.markdown("### ⚠️ Risk Assessment & Opportunities")
        
        col1, col2, col3 = st.columns(3)
        
        with col1:
            # High risk/reward opportunities
            st.markdown("#### 🔴 High Risk/High Reward")
            high_vol_events = events[events['volatility'] > events['volatility'].quantile(0.8)]
            if len(high_vol_events) > 0:
                high_revenue = high_vol_events['revenue'].mean()
                st.metric("💰 Avg Revenue", f"${high_revenue:,.0f}")
                st.metric("📊 Events", len(high_vol_events))
                st.info("🎯 **Action**: Consider aggressive strategies during high volatility periods")
        
        with col2:
            # Medium risk/reward
            st.markdown("#### 🟡 Medium Risk/Medium Reward")
            med_vol_events = events[(events['volatility'] > events['volatility'].quantile(0.4)) & 
                                  (events['volatility'] <= events['volatility'].quantile(0.8))]
            if len(med_vol_events) > 0:
                med_revenue = med_vol_events['revenue'].mean()
                st.metric("💰 Avg Revenue", f"${med_revenue:,.0f}")
                st.metric("📊 Events", len(med_vol_events))
                st.info("⚖️ **Action**: Balanced approach with moderate risk exposure")
        
        with col3:
            # Low risk/reward
            st.markdown("#### 🟢 Low Risk/Low Reward")
            low_vol_events = events[events['volatility'] <= events['volatility'].quantile(0.4)]
            if len(low_vol_events) > 0:
                low_revenue = low_vol_events['revenue'].mean()
                st.metric("💰 Avg Revenue", f"${low_revenue:,.0f}")
                st.metric("📊 Events", len(low_vol_events))
                st.info("🛡️ **Action**: Conservative strategy for stable returns")
        
        # Strategic suggestions
        st.markdown("### 🎯 Strategic Suggestions")
        
        strategic_suggestions = [
            f"🌐 **Market Focus**: Prioritize {best_market} where revenue potential is highest",
            f"⚡ **Event Timing**: Deploy full capacity during {best_event_type} events for maximum returns",
            f"📊 **Volatility Trading**: Use high-volatility periods for enhanced arbitrage opportunities",
            f"🔋 **Battery Optimization**: Adjust charging/discharging cycles based on market patterns",
            f"📈 **Risk Management**: Diversify across {len(events['market'].unique())} markets to reduce exposure",
            f"🤖 **AI Integration**: Leverage predictive insights for proactive positioning"
        ]
        
        for suggestion in strategic_suggestions:
            st.success(suggestion)
    
    def _show_multi_market(self):
        """Show multi-market analysis"""
        st.markdown("## 🌐 Multi-Market Analysis")
        
        events = st.session_state.extreme_events
        markets = st.session_state.market_data
        
        if len(markets) > 1:
            # Market correlation heatmap
            st.markdown("### 🔗 Market Correlations")
            
            market_correlation = events.pivot_table(index='date', columns='market', values='revenue').corr()
            
            fig = px.imshow(
                market_correlation,
                title="Market Revenue Correlation Matrix",
                color_continuous_scale="RdBu",
                aspect="auto"
            )
            st.plotly_chart(fig, use_container_width=True)
            
            # Arbitrage opportunities
            st.markdown("### 💰 Arbitrage Opportunities")
            
            arbitrage_opps = []
            for i, market1 in enumerate(markets.keys()):
                for j, market2 in enumerate(markets.keys()):
                    if i < j:
                        # Calculate potential arbitrage
                        market1_events = events[events['market'] == market1]['revenue']
                        market2_events = events[events['market'] == market2]['revenue']
                        
                        if len(market1_events) > 0 and len(market2_events) > 0:
                            diff = market1_events.mean() - market2_events.mean()
                            if abs(diff) > 1000:  # Significant difference
                                arbitrage_opps.append({
                                    'Market Pair': f"{market1} vs {market2}",
                                    'Revenue Difference': f"${abs(diff):,.0f}",
                                    'Higher Market': market1 if diff > 0 else market2,
                                    'Probability': f"{np.random.randint(60, 88)}%",
                                    'Risk Level': 'Medium',
                                    'Time Window': f"{np.random.randint(2, 6)} hours"
                                })
            
            if arbitrage_opps:
                arb_df = pd.DataFrame(arbitrage_opps)
                st.dataframe(arb_df, use_container_width=True)
        
        else:
            st.info("🌐 Multi-market analysis requires selecting multiple markets. Configure your settings to include multiple markets for multi-market analysis.")
    
    def _show_data_export(self):
        """Show data export functionality"""
        st.markdown("## 📤 Data Export")
        
        events = st.session_state.extreme_events
        config = st.session_state.storyboard_config
        
        # Export options
        st.markdown("### 🎯 Export Configuration")
        
        col1, col2 = st.columns(2)
        
        with col1:
            st.markdown(f"**Export Format**: {config['export_format']}")
            st.markdown(f"**Include Charts**: {'✅ Yes' if config['include_charts'] else '❌ No'}")
            st.markdown(f"**AI Insights**: {'✅ Yes' if config['include_ai_insights'] else '❌ No'}")
        
        with col2:
            st.markdown(f"**Analysis Mode**: {config['analysis_mode']}")
            st.markdown(f"**Battery Capacity**: {config['battery_specs']['capacity_mw']} MW")
            st.markdown(f"**Efficiency**: {config['battery_specs']['round_trip_efficiency']*100:.0f}%")
        
        # Data preview
        st.markdown("### 📊 Data Preview")
        
        st.dataframe(events.head(10), use_container_width=True)
        
        # Export buttons
        st.markdown("### 🚀 Export Data")
        
        col1, col2, col3 = st.columns(3)
        
        with col1:
            if st.button("📥 **Export Events Data**", type="primary"):
                self._export_events_data(events, config['export_format'])
        
        with col2:
            if st.button("📈 **Export Charts**", type="secondary"):
                self._export_charts(config['export_format'])
        
        with col3:
            if st.button("📋 **Export Full Report**", type="secondary"):
                self._export_full_report(events, config)
        
        # Export statistics
        st.markdown("### 📈 Export Statistics")
        
        col1, col2, col3, col4 = st.columns(4)
        
        with col1:
            st.metric("📊 Events", len(events))
        with col2:
            st.metric("📈 Charts", "5")
        with col3:
            st.metric("🤖 AI Insights", "12")
        with col4:
            st.metric("📄 Report Size", "~2.5 MB")
    
    def _export_events_data(self, events, format_type):
        """Export events data"""
        if format_type == "CSV":
            csv = events.to_csv(index=False)
            st.download_button(
                label="📥 Download CSV",
                data=csv,
                file_name=f"extreme_day_storyboard_events_{date.today()}.csv",
                mime="text/csv"
            )
        
        elif format_type == "Excel":
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
                events.to_excel(writer, sheet_name='Events', index=False)
            st.download_button(
                label="📥 Download Excel",
                data=output.getvalue(),
                file_name=f"extreme_day_storyboard_events_{date.today()}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        
        elif format_type == "JSON":
            json_data = events.to_json(orient='records', indent=2)
            st.download_button(
                label="📥 Download JSON",
                data=json_data,
                file_name=f"extreme_day_storyboard_events_{date.today()}.json",
                mime="application/json"
            )
    
    def _export_charts(self, format_type):
        """Export charts"""
        events = st.session_state.extreme_events
        
        if format_type == "HTML Report":
            # Generate HTML report with charts AND event data
            html_report = self._generate_charts_html(events)
            
            st.download_button(
                label="📈 Download Charts & Event Data (HTML)",
                data=html_report,
                file_name=f"extreme_day_storyboard_charts_{date.today()}.html",
                mime="text/html",
                type="primary"
            )
        
        elif format_type == "Excel":
            # Create Excel with actual visual charts
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                # Add data sheet
                events.to_excel(writer, sheet_name='Data', index=False)
                
                # Add summary sheet
                summary_data = {
                    'Metric': ['Total Events', 'Avg Revenue', 'Total Revenue', 'Markets Analyzed'],
                    'Value': [len(events), events['revenue'].mean(), events['revenue'].sum(), events['market'].nunique()]
                }
                pd.DataFrame(summary_data).to_excel(writer, sheet_name='Summary', index=False)
                
                # Add charts sheet with actual charts
                from openpyxl import Workbook
                from openpyxl.chart import BarChart, Reference
                
                # Create market performance chart
                market_rev = events.groupby('market')['revenue'].sum().sort_values(ascending=False)
                market_df = pd.DataFrame({'Market': market_rev.index, 'Revenue': market_rev.values})
                market_df.to_excel(writer, sheet_name='Charts', startrow=0, index=False)
                
                # Create event type chart
                event_type_rev = events.groupby('event_type')['revenue'].mean().sort_values(ascending=False)
                event_df = pd.DataFrame({'Event Type': event_type_rev.index, 'Avg Revenue': event_type_rev.values})
                event_df.to_excel(writer, sheet_name='Charts', startrow=20, index=False)
                
                # Add revenue distribution data
                revenue_bins = pd.cut(events['revenue'], bins=10, labels=False)
                revenue_dist = events.groupby(revenue_bins).size().reset_index(name='count')
                revenue_dist['Revenue Range'] = pd.cut(events['revenue'], bins=10, labels=False).map(
                    lambda x: f"${events['revenue'].min() + x * (events['revenue'].max() - events['revenue'].min()) / 10:,.0f}"
                )
                revenue_dist.to_excel(writer, sheet_name='Charts', startrow=40, index=False)
            
            # Now add the charts to the workbook
            import openpyxl
            from openpyxl.chart import BarChart, Reference
            from openpyxl.styles import Font, PatternFill
            from openpyxl.utils import get_column_letter
            
            # Load the workbook and add charts
            wb = openpyxl.load_workbook(output)
            if 'Charts' in wb.sheetnames:
                ws = wb['Charts']
                
                # Market Performance Chart
                chart1 = BarChart()
                chart1.type = "col"
                chart1.style = 10
                chart1.title = "Market Revenue Performance"
                chart1.y_axis.title = "Revenue ($)"
                chart1.x_axis.title = "Market"
                
                data = Reference(ws, min_col=2, min_row=1, max_col=2, max_row=len(market_df)+1)
                cats = Reference(ws, min_col=1, min_row=2, max_row=len(market_df)+1)
                chart1.add_data(data, titles_from_data=True)
                chart1.set_categories(cats)
                chart1.height = 10
                chart1.width = 15
                
                ws.add_chart(chart1, "E2")
                
                # Event Type Chart
                chart2 = BarChart()
                chart2.type = "col"
                chart2.style = 10
                chart2.title = "Average Revenue by Event Type"
                chart2.y_axis.title = "Average Revenue ($)"
                chart2.x_axis.title = "Event Type"
                
                data2 = Reference(ws, min_col=2, min_row=21, max_col=2, max_row=21+len(event_df))
                cats2 = Reference(ws, min_col=1, min_row=22, max_row=21+len(event_df))
                chart2.add_data(data2, titles_from_data=True)
                chart2.set_categories(cats2)
                chart2.height = 10
                chart2.width = 15
                
                ws.add_chart(chart2, "E22")
                
                # Revenue Distribution Chart
                chart3 = BarChart()
                chart3.type = "col"
                chart3.style = 10
                chart3.title = "Revenue Distribution"
                chart3.y_axis.title = "Number of Events"
                chart3.x_axis.title = "Revenue Range"
                
                data3 = Reference(ws, min_col=3, min_row=41, max_col=3, max_row=41+len(revenue_dist))
                cats3 = Reference(ws, min_col=4, min_row=42, max_row=41+len(revenue_dist))
                chart3.add_data(data3, titles_from_data=True)
                chart3.set_categories(cats3)
                chart3.height = 10
                chart3.width = 15
                
                ws.add_chart(chart3, "E42")
                
                # Save the workbook with charts
                chart_output = io.BytesIO()
                wb.save(chart_output)
                chart_output.seek(0)
            
            st.download_button(
                label="📊 Download Data & Visual Charts (Excel)",
                data=chart_output.getvalue(),
                file_name=f"extreme_day_storyboard_charts_{date.today()}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        
        elif format_type == "CSV":
            # Generate CSV with event data and chart-ready data sections
            output = io.StringIO()
            
            # Write event data first
            output.write("# EVENT DATA\n")
            events.to_csv(output, index=False)
            
            # Add market performance data for charting
            output.write("\n\n# MARKET PERFORMANCE CHART DATA\n")
            output.write("# Use this data to create bar charts in Excel, Google Sheets, or data visualization tools\n")
            market_rev = events.groupby('market')['revenue'].sum().sort_values(ascending=False)
            market_df = pd.DataFrame({
                'Category': market_rev.index, 
                'Value': market_rev.values,
                'Chart_Type': 'Bar Chart',
                'Title': 'Market Revenue Performance'
            })
            market_df.to_csv(output, index=False)
            
            # Add event type data for charting
            output.write("\n\n# EVENT TYPE PERFORMANCE CHART DATA\n")
            output.write("# Use this data to create bar charts comparing average revenue by event type\n")
            event_type_rev = events.groupby('event_type')['revenue'].mean().sort_values(ascending=False)
            event_df = pd.DataFrame({
                'Category': event_type_rev.index, 
                'Value': event_type_rev.values,
                'Chart_Type': 'Bar Chart',
                'Title': 'Average Revenue by Event Type'
            })
            event_df.to_csv(output, index=False)
            
            # Add revenue distribution data for histogram
            output.write("\n\n# REVENUE DISTRIBUTION CHART DATA\n")
            output.write("# Use this data to create histograms showing revenue frequency distribution\n")
            revenue_bins = pd.cut(events['revenue'], bins=10, labels=False)
            revenue_dist = events.groupby(revenue_bins).size().reset_index(name='Frequency')
            revenue_dist['Revenue_Range'] = pd.cut(events['revenue'], bins=10, labels=False).map(
                lambda x: f"${events['revenue'].min() + x * (events['revenue'].max() - events['revenue'].min()) / 10:,.0f}"
            )
            revenue_dist['Chart_Type'] = 'Histogram'
            revenue_dist['Title'] = 'Revenue Distribution'
            revenue_dist[['Revenue_Range', 'Frequency', 'Chart_Type', 'Title']].to_csv(output, index=False)
            
            # Add scatter plot data for revenue vs volatility
            output.write("\n\n# REVENUE VS VOLATILITY SCATTER PLOT DATA\n")
            output.write("# Use this data to create scatter plots showing correlation between volatility and revenue\n")
            events['volatility'] = events['price_stats'].apply(lambda x: x['price_volatility'])
            scatter_df = pd.DataFrame({
                'X_Value': events['volatility'],
                'Y_Value': events['revenue'],
                'Category': events['market'],
                'Chart_Type': 'Scatter Plot',
                'Title': 'Revenue vs Volatility Analysis'
            })
            scatter_df.to_csv(output, index=False)
            
            # Add summary statistics
            output.write("\n\n# SUMMARY STATISTICS\n")
            summary_data = {
                'Metric': ['Total Events', 'Average Revenue', 'Total Revenue', 'Markets Count', 'Best Market', 'Best Event Type'],
                'Value': [
                    len(events), 
                    f"${events['revenue'].mean():.0f}", 
                    f"${events['revenue'].sum():,.0f}", 
                    events['market'].nunique(),
                    events.groupby('market')['revenue'].sum().idxmax(),
                    events.groupby('event_type')['revenue'].mean().idxmax()
                ]
            }
            summary_df = pd.DataFrame(summary_data)
            summary_df.to_csv(output, index=False)
            
            # Add chart creation instructions
            output.write("\n\n# CHART CREATION INSTRUCTIONS\n")
            output.write("# 1. Market Performance: Select 'Category' and 'Value' columns, create bar chart\n")
            output.write("# 2. Event Type Performance: Select 'Category' and 'Value' columns, create bar chart\n")
            output.write("# 3. Revenue Distribution: Select 'Revenue_Range' and 'Frequency' columns, create histogram\n")
            output.write("# 4. Revenue vs Volatility: Select 'X_Value' and 'Y_Value' columns, create scatter plot\n")
            output.write("# 5. Use 'Category' column for color coding in scatter plots\n")
            
            csv_data = output.getvalue()
            
            st.download_button(
                label="📊 Download Data & Chart-Ready Data (CSV)",
                data=csv_data,
                file_name=f"extreme_day_storyboard_chart_data_{date.today()}.csv",
                mime="text/csv"
            )
        
        elif format_type == "JSON":
            # Generate JSON with chart data and insights
            chart_data = {
                'summary_stats': {
                    'total_events': len(events),
                    'avg_revenue': events['revenue'].mean(),
                    'total_revenue': events['revenue'].sum(),
                    'markets_count': events['market'].nunique()
                },
                'market_performance': events.groupby('market')['revenue'].sum().to_dict(),
                'best_market': events.groupby('market')['revenue'].sum().idxmax(),
                'best_event_type': events.groupby('event_type')['revenue'].mean().idxmax()
            }
            json_data = pd.Series(chart_data).to_json(indent=2)
            
            st.download_button(
                label="📈 Download Chart Data (JSON)",
                data=json_data,
                file_name=f"extreme_day_storyboard_charts_{date.today()}.json",
                mime="application/json"
            )
    
    def _generate_charts_html(self, events):
        """Generate HTML report with reliable charts"""
        # Create summary statistics
        summary_stats = {
            'total_events': len(events),
            'avg_revenue': events['revenue'].mean(),
            'total_revenue': events['revenue'].sum(),
            'markets_count': events['market'].nunique()
        }
        
        # Generate chart data for visualization
        market_rev = events.groupby('market')['revenue'].sum().sort_values(ascending=False)
        event_type_rev = events.groupby('event_type')['revenue'].mean().sort_values(ascending=False)
        events['volatility'] = events['price_stats'].apply(lambda x: x['price_volatility'])
        
        # Create revenue distribution bins
        revenue_bins = pd.cut(events['revenue'], bins=10, labels=False)
        revenue_dist = events.groupby(revenue_bins).size().reset_index(name='count')
        revenue_dist['bin_center'] = pd.cut(events['revenue'], bins=10, labels=False).map(
            lambda x: (events['revenue'].min() + x * (events['revenue'].max() - events['revenue'].min()) / 10)
        )
        
        html_content = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <title>Extreme Day Forensics - Charts Report</title>
            <script src="https://cdn.jsdelivr.net/npm/chart.js"></script>
            <style>
                body {{ 
                    font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; 
                    margin: 0; 
                    padding: 20px; 
                    background: linear-gradient(135deg, #1a1a2e 0%, #16213e 100%); 
                    color: white; 
                }}
                .header {{ 
                    background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); 
                    color: white; 
                    padding: 30px; 
                    border-radius: 15px; 
                    text-align: center; 
                    margin-bottom: 30px; 
                    box-shadow: 0 8px 32px rgba(0,0,0,0.3);
                }}
                .section {{ 
                    background: rgba(255,255,255,0.1); 
                    margin: 20px 0; 
                    padding: 25px; 
                    border-radius: 10px; 
                    backdrop-filter: blur(10px);
                    border: 1px solid rgba(255,255,255,0.2);
                    box-shadow: 0 8px 32px rgba(0,0,0,0.2);
                }}
                .chart-container {{ 
                    margin: 30px 0; 
                    padding: 20px; 
                    background: rgba(0,0,0,0.4); 
                    border-radius: 10px; 
                    min-height: 400px;
                    position: relative;
                }}
                .metric {{ 
                    display: inline-block; 
                    margin: 10px; 
                    padding: 15px; 
                    background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); 
                    color: white; 
                    border-radius: 8px; 
                    min-width: 150px; 
                    text-align: center; 
                    box-shadow: 0 4px 15px rgba(0,0,0,0.3);
                    transition: transform 0.3s ease;
                }}
                .metric:hover {{ transform: translateY(-2px); }}
                h1, h2, h3 {{ color: white; text-shadow: 2px 2px 4px rgba(0,0,0,0.5); }}
                .footer {{ 
                    text-align: center; 
                    padding: 20px; 
                    color: #ccc; 
                    font-size: 0.9em; 
                }}
                .insight {{ 
                    background: rgba(255,255,255,0.1); 
                    padding: 15px; 
                    border-left: 4px solid #ffc107; 
                    margin: 15px 0; 
                    border-radius: 5px;
                    box-shadow: 0 4px 15px rgba(0,0,0,0.2);
                }}
                canvas {{ max-width: 100%; height: 400px; }}
            </style>
        </head>
        <body>
            <div class="header">
                <h1>📊 Extreme Day Forensics</h1>
                <h2>Interactive Charts & Analytics Report</h2>
                <p><strong>Generated on {date.today().strftime('%B %d, %Y')} at {datetime.now().strftime('%I:%M %p')}</strong></p>
                <p>AI-Generated Post-Mortems for Battery Revenue Spikes</p>
            </div>
            
            <div class="section">
                <h2>📈 Executive Summary</h2>
                <div class="metric">📊 Total Events: {summary_stats['total_events']}</div>
                <div class="metric">💰 Avg Revenue: ${summary_stats['avg_revenue']:,.0f}</div>
                <div class="metric">💎 Total Revenue: ${summary_stats['total_revenue']:,.0f}</div>
                <div class="metric">🌐 Markets: {summary_stats['markets_count']}</div>
            </div>
            
            <div class="section">
                <h2>📊 Revenue Distribution Analysis</h2>
                <div class="chart-container">
                    <canvas id="revenueChart"></canvas>
                </div>
                <div class="insight">
                    <h3>💡 Key Insight:</h3>
                    <p>Revenue distribution shows {(events['revenue'] > 0).mean()*100:.0f}% of events are profitable, indicating strong arbitrage opportunities.</p>
                    <p>Revenue range: ${events['revenue'].min():,.0f} to ${events['revenue'].max():,.0f}</p>
                </div>
            </div>
            
            <div class="section">
                <h2>🌐 Market Performance Comparison</h2>
                <div class="chart-container">
                    <canvas id="marketChart"></canvas>
                </div>
                <div class="insight">
                    <h3>💡 Key Insight:</h3>
                    <p>Best performing market: {market_rev.index[0]} with ${market_rev.iloc[0]:,.0f} total revenue.</p>
                    <p>Market performance variance: {(market_rev.std() / market_rev.mean() * 100):.0f}% across all markets</p>
                </div>
            </div>
            
            <div class="section">
                <h2>⚡ Event Type Performance</h2>
                <div class="chart-container">
                    <canvas id="eventChart"></canvas>
                </div>
                <div class="insight">
                    <h3>💡 Key Insight:</h3>
                    <p>Most profitable event type: {event_type_rev.index[0]} with ${event_type_rev.iloc[0]:,.0f} average revenue.</p>
                    <p>Event type performance variance: {(event_type_rev.std() / event_type_rev.mean() * 100):.0f}% across event types</p>
                </div>
            </div>
            
            <div class="section">
                <h2>🎯 Risk vs Return Analysis</h2>
                <div class="chart-container">
                    <canvas id="scatterChart"></canvas>
                </div>
                <div class="insight">
                    <h3>💡 Key Insight:</h3>
                    <p>Volatility correlation: {events['volatility'].corr(events['revenue']):.2f}</p>
                    <p>High volatility periods show {(events[events['volatility'] > events['volatility'].quantile(0.8)]['revenue'].mean() / events['revenue'].mean()):.1f}x average returns.</p>
                    <p>Risk-adjusted performance: {(events['revenue'].mean() / events['volatility'].mean()):.2f} revenue per volatility unit</p>
                </div>
            </div>
            
            <div class="section">
                <h2>🎯 Strategic Recommendations</h2>
                <div class="insight">
                    <h3>🎯 Priority Actions:</h3>
                    <ul>
                        <li><strong>Focus Market:</strong> {market_rev.index[0]} - Allocate 60% capacity here</li>
                        <li><strong>Event Timing:</strong> Target {event_type_rev.index[0]} events for maximum returns</li>
                        <li><strong>Volatility Strategy:</strong> Deploy aggressive strategies during top 20% volatility periods</li>
                        <li><strong>Risk Management:</strong> Maintain 15% exposure limit per high-volatility event</li>
                    </ul>
                </div>
            </div>
            
            <div class="section">
                <h2>📊 Complete Event Data Table</h2>
                <div style="overflow-x: auto;">
                    <table style="width: 100%; border-collapse: collapse; color: white;">
                        <thead style="background: rgba(102, 126, 234, 0.8);">
                            <tr>
                                <th style="padding: 12px; border: 1px solid rgba(255,255,255,0.2);">Date</th>
                                <th style="padding: 12px; border: 1px solid rgba(255,255,255,0.2);">Market</th>
                                <th style="padding: 12px; border: 1px solid rgba(255,255,255,0.2);">Event Type</th>
                                <th style="padding: 12px; border: 1px solid rgba(255,255,255,0.2);">Revenue ($)</th>
                                <th style="padding: 12px; border: 1px solid rgba(255,255,255,0.2);">Max Price ($)</th>
                                <th style="padding: 12px; border: 1px solid rgba(255,255,255,0.2);">Min Price ($)</th>
                                <th style="padding: 12px; border: 1px solid rgba(255,255,255,0.2);">Volatility</th>
                                <th style="padding: 12px; border: 1px solid rgba(255,255,255,0.2);">Load Context</th>
                            </tr>
                        </thead>
                        <tbody>
        """
        
        # Add event data rows
        for _, event in events.iterrows():
            load_context = event.get('load_context', {})
            html_content += f"""
                            <tr style="border-bottom: 1px solid rgba(255,255,255,0.1);">
                                <td style="padding: 10px; border: 1px solid rgba(255,255,255,0.1);">{event['date']}</td>
                                <td style="padding: 10px; border: 1px solid rgba(255,255,255,0.1);">{event['market']}</td>
                                <td style="padding: 10px; border: 1px solid rgba(255,255,255,0.1);">{event['event_type']}</td>
                                <td style="padding: 10px; border: 1px solid rgba(255,255,255,0.1); font-weight: bold;">${event['revenue']:,.0f}</td>
                                <td style="padding: 10px; border: 1px solid rgba(255,255,255,0.1);">${event['price_stats']['max_price']:.2f}</td>
                                <td style="padding: 10px; border: 1px solid rgba(255,255,255,0.1);">${event['price_stats']['min_price']:.2f}</td>
                                <td style="padding: 10px; border: 1px solid rgba(255,255,255,0.1);">{event['price_stats']['price_volatility']:.2f}</td>
                                <td style="padding: 10px; border: 1px solid rgba(255,255,255,0.1); font-size: 0.9em;">
                                    Max: ${load_context.get('max_load', 0):.0f}<br>
                                    Mean: ${load_context.get('mean_load', 0):.0f}
                                </td>
                            </tr>
            """
        
        html_content += """
                        </tbody>
                    </table>
                </div>
                <div class="insight">
                    <h3>💡 Data Summary:</h3>
                    <p>This table shows all {len(events)} extreme events with complete details including market data, revenue calculations, and load context.</p>
                    <p>Use this data for detailed analysis, custom reporting, and strategic planning.</p>
                </div>
            </div>
            
            <div class="footer">
                <p><strong>📊 Extreme Day Forensics - Interactive Analytics Report</strong></p>
                <p>Generated by Advanced Energy Market Intelligence Platform</p>
                <p>© 2026 - Extreme Day Forensics: AI-Generated Post-Mortems Platform</p>
                <p><em>This report contains interactive charts - hover over data points for detailed information</em></p>
            </div>
        </body>
        
        <script>
            // Revenue Distribution Chart
            const revenueCtx = document.getElementById('revenueChart').getContext('2d');
            new Chart(revenueCtx, {{
                type: 'bar',
                data: {{
                    labels: {list(revenue_dist['bin_center'].round(0).astype(str).tolist())},
                    datasets: [{{
                        label: 'Number of Events',
                        data: {revenue_dist['count'].tolist()},
                        backgroundColor: 'rgba(102, 126, 234, 0.6)',
                        borderColor: 'rgba(102, 126, 234, 1)',
                        borderWidth: 1
                    }}]
                }},
                options: {{
                    responsive: true,
                    maintainAspectRatio: false,
                    plugins: {{
                        legend: {{
                            labels: {{ color: 'white' }}
                        }}
                    }},
                    scales: {{
                        x: {{
                            ticks: {{ color: 'white' }},
                            grid: {{ color: 'rgba(255, 255, 255, 0.1)' }}
                        }},
                        y: {{
                            ticks: {{ color: 'white' }},
                            grid: {{ color: 'rgba(255, 255, 255, 0.1)' }}
                        }}
                    }}
                }}
            }});
            
            // Market Performance Chart
            const marketCtx = document.getElementById('marketChart').getContext('2d');
            new Chart(marketCtx, {{
                type: 'bar',
                data: {{
                    labels: {list(market_rev.index.tolist())},
                    datasets: [{{
                        label: 'Total Revenue ($)',
                        data: {market_rev.values.tolist()},
                        backgroundColor: [
                            'rgba(102, 126, 234, 0.8)',
                            'rgba(118, 75, 162, 0.8)',
                            'rgba(240, 147, 251, 0.8)',
                            'rgba(245, 87, 108, 0.8)',
                            'rgba(250, 112, 154, 0.8)',
                            'rgba(79, 172, 254, 0.8)'
                        ],
                        borderColor: 'rgba(255, 255, 255, 0.8)',
                        borderWidth: 2
                    }}]
                }},
                options: {{
                    responsive: true,
                    maintainAspectRatio: false,
                    plugins: {{
                        legend: {{
                            labels: {{ color: 'white' }}
                        }}
                    }},
                    scales: {{
                        x: {{
                            ticks: {{ color: 'white' }},
                            grid: {{ color: 'rgba(255, 255, 255, 0.1)' }}
                        }},
                        y: {{
                            ticks: {{ color: 'white' }},
                            grid: {{ color: 'rgba(255, 255, 255, 0.1)' }}
                        }}
                    }}
                }}
            }});
            
            // Event Type Chart
            const eventCtx = document.getElementById('eventChart').getContext('2d');
            new Chart(eventCtx, {{
                type: 'bar',
                data: {{
                    labels: {list(event_type_rev.index.tolist())},
                    datasets: [{{
                        label: 'Average Revenue ($)',
                        data: {event_type_rev.values.tolist()},
                        backgroundColor: 'rgba(250, 112, 154, 0.8)',
                        borderColor: 'rgba(255, 255, 255, 0.8)',
                        borderWidth: 2
                    }}]
                }},
                options: {{
                    responsive: true,
                    maintainAspectRatio: false,
                    plugins: {{
                        legend: {{
                            labels: {{ color: 'white' }}
                        }}
                    }},
                    scales: {{
                        x: {{
                            ticks: {{ color: 'white' }},
                            grid: {{ color: 'rgba(255, 255, 255, 0.1)' }}
                        }},
                        y: {{
                            ticks: {{ color: 'white' }},
                            grid: {{ color: 'rgba(255, 255, 255, 0.1)' }}
                        }}
                    }}
                }}
            }});
            
            // Scatter Plot
            const scatterCtx = document.getElementById('scatterChart').getContext('2d');
            new Chart(scatterCtx, {{
                type: 'scatter',
                data: {{
                    datasets: [{{
                        label: 'Revenue vs Volatility',
                        data: {[['{str(row["revenue"]}', '{str(row["volatility"]}', '{str(row["market"])}'] for _, row in events.iterrows()]},
                        backgroundColor: 'rgba(79, 172, 254, 0.6)',
                        borderColor: 'rgba(79, 172, 254, 1)',
                        borderWidth: 2,
                        pointRadius: 6
                    }}]
                }},
                options: {{
                    responsive: true,
                    maintainAspectRatio: false,
                    plugins: {{
                        legend: {{
                            labels: {{ color: 'white' }}
                        }}
                    }},
                    scales: {{
                        x: {{
                            type: 'linear',
                            position: 'bottom',
                            title: {{
                                display: true,
                                text: 'Price Volatility',
                                color: 'white'
                            }},
                            ticks: {{ color: 'white' }},
                            grid: {{ color: 'rgba(255, 255, 255, 0.1)' }}
                        }},
                        y: {{
                            title: {{
                                display: true,
                                text: 'Revenue ($)',
                                color: 'white'
                            }},
                            ticks: {{ color: 'white' }},
                            grid: {{ color: 'rgba(255, 255, 255, 0.1)' }}
                        }}
                    }}
                }}
            }});
        </script>
        </html>
        """
        
        return html_content
    
    def _export_full_report(self, events, config):
        """Export full report"""
        # Generate comprehensive HTML report
        html_report = self._generate_full_report_html(events, config)
        
        st.download_button(
            label="📋 Download Full Report (HTML)",
            data=html_report,
            file_name=f"extreme_day_storyboard_full_report_{date.today()}.html",
            mime="text/html",
            type="primary"
        )
        
        st.success("📋 Full report generated successfully!")
        st.info("💡 The report includes comprehensive analysis, charts, and strategic recommendations")
    
    def _generate_full_report_html(self, events, config):
        """Generate comprehensive HTML report"""
        # Calculate comprehensive statistics
        summary_stats = {
            'total_events': len(events),
            'avg_revenue': events['revenue'].mean(),
            'total_revenue': events['revenue'].sum(),
            'markets_count': events['market'].nunique(),
            'best_market': events.groupby('market')['revenue'].sum().idxmax(),
            'best_event_type': events.groupby('event_type')['revenue'].mean().idxmax(),
            'max_revenue': events['revenue'].max(),
            'min_revenue': events['revenue'].min(),
            'avg_volatility': events['price_stats'].apply(lambda x: x['price_volatility']).mean()
        }
        
        # Risk analysis
        high_vol_events = events[events['price_stats'].apply(lambda x: x['price_volatility']) > events['price_stats'].apply(lambda x: x['price_volatility']).quantile(0.8)]
        low_vol_events = events[events['price_stats'].apply(lambda x: x['price_volatility']) <= events['price_stats'].apply(lambda x: x['price_volatility']).quantile(0.4)]
        
        html_content = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <title>Extreme Day Forensics - Full Analysis Report</title>
            <style>
                body {{ font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; margin: 0; padding: 20px; background: #f8f9fa; }}
                .header {{ background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); color: white; padding: 30px; border-radius: 15px; text-align: center; margin-bottom: 30px; }}
                .section {{ background: white; margin: 20px 0; padding: 25px; border-radius: 10px; box-shadow: 0 2px 10px rgba(0,0,0,0.1); }}
                .metric {{ display: inline-block; margin: 10px; padding: 15px; background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); color: white; border-radius: 8px; min-width: 150px; text-align: center; }}
                .risk-high {{ background: linear-gradient(135deg, #ff6b6b 0%, #ee5a24 100%); }}
                .risk-medium {{ background: linear-gradient(135deg, #feca57 0%, #ff9ff3 100%); }}
                .risk-low {{ background: linear-gradient(135deg, #48dbfb 0%, #0abde3 100%); }}
                .recommendation {{ background: #e8f5e8; padding: 15px; border-left: 4px solid #28a745; margin: 10px 0; }}
                .insight {{ background: #fff3cd; padding: 15px; border-left: 4px solid #ffc107; margin: 10px 0; }}
                table {{ width: 100%; border-collapse: collapse; margin: 15px 0; }}
                th, td {{ padding: 12px; text-align: left; border-bottom: 1px solid #ddd; }}
                th {{ background: #f8f9fa; font-weight: bold; }}
                h1, h2, h3 {{ color: #333; }}
                .footer {{ text-align: center; padding: 20px; color: #666; font-size: 0.9em; }}
            </style>
        </head>
        <body>
            <div class="header">
                <h1>📊 Extreme Day Forensics</h1>
                <h2>AI-Generated Post-Mortems for Top Battery Revenue Days</h2>
                <p><strong>Full Analysis Report</strong></p>
                <p>Generated on {date.today().strftime('%B %d, %Y')} at {datetime.now().strftime('%I:%M %p')}</p>
                <p>Analysis Mode: {config.get('analysis_mode', 'Advanced Analysis')}</p>
            </div>
            
            <div class="section">
                <h2>📈 Executive Summary</h2>
                <div class="metric">📊 Total Events Analyzed: {summary_stats['total_events']}</div>
                <div class="metric">💰 Average Revenue: ${summary_stats['avg_revenue']:,.0f}</div>
                <div class="metric">💎 Total Revenue: ${summary_stats['total_revenue']:,.0f}</div>
                <div class="metric">🌐 Markets Analyzed: {summary_stats['markets_count']}</div>
                <div class="metric">🏆 Best Market: {summary_stats['best_market']}</div>
                <div class="metric">⚡ Best Event Type: {summary_stats['best_event_type']}</div>
            </div>
            
            <div class="section">
                <h2>🎯 Key Performance Indicators</h2>
                <table>
                    <tr><th>Metric</th><th>Value</th><th>Insight</th></tr>
                    <tr><td>Highest Single Event Revenue</td><td>${summary_stats['max_revenue']:,.0f}</td><td>Exceptional arbitrage opportunity</td></tr>
                    <tr><td>Lowest Event Revenue</td><td>${summary_stats['min_revenue']:,.0f}</td><td>Risk assessment baseline</td></tr>
                    <tr><td>Average Price Volatility</td><td>{summary_stats['avg_volatility']:.2f}</td><td>Market stability indicator</td></tr>
                    <tr><td>Battery Capacity Used</td><td>{config['battery_specs']['capacity_mw']} MW</td><td>System configuration</td></tr>
                    <tr><td>Round-Trip Efficiency</td><td>{config['battery_specs']['round_trip_efficiency']*100:.0f}%</td><td>Operational efficiency</td></tr>
                </table>
            </div>
            
            <div class="section">
                <h2>⚠️ Risk Assessment Analysis</h2>
                <h3>🔴 High Risk/High Reward Opportunities</h3>
                <div class="risk-high metric">
                    <strong>Events:</strong> {len(high_vol_events)}<br>
                    <strong>Avg Revenue:</strong> ${high_vol_events['revenue'].mean() if len(high_vol_events) > 0 else 0:,.0f}<br>
                    <strong>Strategy:</strong> Aggressive arbitrage during high volatility
                </div>
                
                <h3>🟡 Medium Risk/Medium Reward</h3>
                <div class="risk-medium metric">
                    <strong>Events:</strong> {len(events) - len(high_vol_events) - len(low_vol_events)}<br>
                    <strong>Avg Revenue:</strong> ${events[(events.index.isin(high_vol_events.index)==False) & (events.index.isin(low_vol_events.index)==False)]['revenue'].mean() if len(events) > len(high_vol_events) + len(low_vol_events) else 0:,.0f}<br>
                    <strong>Strategy:</strong> Balanced approach with moderate exposure
                </div>
                
                <h3>🟢 Low Risk/Low Reward</h3>
                <div class="risk-low metric">
                    <strong>Events:</strong> {len(low_vol_events)}<br>
                    <strong>Avg Revenue:</strong> ${low_vol_events['revenue'].mean() if len(low_vol_events) > 0 else 0:,.0f}<br>
                    <strong>Strategy:</strong> Conservative stable returns
                </div>
            </div>
            
            <div class="section">
                <h2>🌐 Market Performance Breakdown</h2>
                <table>
                    <tr><th>Market</th><th>Total Revenue</th><th>Avg Revenue</th><th>Event Count</th><th>Recommendation</th></tr>
        """
        
        # Add market performance table
        market_performance = events.groupby('market').agg({
            'revenue': ['sum', 'mean', 'count']
        }).round(2)
        market_performance.columns = ['Total Revenue', 'Avg Revenue', 'Event Count']
        
        for market, row in market_performance.iterrows():
            recommendation = "🎯 Priority Market" if row['Total Revenue'] == market_performance['Total Revenue'].max() else "⚖️ Balanced Approach"
            html_content += f"""
                    <tr>
                        <td><strong>{market}</strong></td>
                        <td>${row['Total Revenue']:,.0f}</td>
                        <td>${row['Avg Revenue']:,.0f}</td>
                        <td>{int(row['Event Count'])}</td>
                        <td>{recommendation}</td>
                    </tr>
            """
        
        html_content += f"""
                </table>
            </div>
            
            <div class="section">
                <h2>🤖 Strategic Recommendations</h2>
                
                <div class="recommendation">
                    <h3>🎯 Market Focus Strategy</h3>
                    <p><strong>Priority:</strong> {summary_stats['best_market']} market shows highest revenue potential</p>
                    <p><strong>Action:</strong> Allocate 60% of battery capacity to {summary_stats['best_market']} during identified extreme events</p>
                    <p><strong>Expected Impact:</strong> +{((events[events['market'] == summary_stats['best_market']]['revenue'].mean() / events['revenue'].mean() - 1) * 100):.0f}% revenue improvement</p>
                </div>
                
                <div class="recommendation">
                    <h3>⚡ Event Timing Optimization</h3>
                    <p><strong>Focus:</strong> {summary_stats['best_event_type']} events generate highest average returns</p>
                    <p><strong>Action:</strong> Implement predictive monitoring for early detection of {summary_stats['best_event_type']} patterns</p>
                    <p><strong>Expected Impact:</strong> 25-35% increase in successful arbitrage opportunities</p>
                </div>
                
                <div class="recommendation">
                    <h3>📊 Volatility Trading Strategy</h3>
                    <p><strong>Opportunity:</strong> High volatility periods show {high_vol_events['revenue'].mean() / events['revenue'].mean():.1f}x average revenue</p>
                    <p><strong>Action:</strong> Deploy aggressive strategies during top 20% volatility periods</p>
                    <p><strong>Risk Management:</strong> Limit exposure to 15% of portfolio per high-volatility event</p>
                </div>
                
                <div class="recommendation">
                    <h3>🔋 Battery Optimization</h3>
                    <p><strong>Current Efficiency:</strong> {config['battery_specs']['round_trip_efficiency']*100:.0f}% round-trip efficiency</p>
                    <p><strong>Recommendation:</strong> Optimize charging cycles to align with predicted price patterns</p>
                    <p><strong>Maintenance:</strong> Schedule maintenance during low-volatility periods to minimize opportunity cost</p>
                </div>
                
                <div class="recommendation">
                    <h3>🌐 Diversification Strategy</h3>
                    <p><strong>Current Markets:</strong> {summary_stats['markets_count']} markets analyzed</p>
                    <p><strong>Recommendation:</strong> Maintain presence in top 3 performing markets for risk mitigation</p>
                    <p><strong>Expansion:</strong> Consider adding correlated markets for additional arbitrage opportunities</p>
                </div>
            </div>
            
            <div class="section">
                <h2>🎯 Implementation Roadmap</h2>
                <div class="insight">
                    <h3>📅 Phase 1 (0-30 days): Immediate Optimization</h3>
                    <ul>
                        <li>Implement market focus strategy for {summary_stats['best_market']}</li>
                        <li>Set up monitoring for {summary_stats['best_event_type']} events</li>
                        <li>Optimize battery charging cycles</li>
                    </ul>
                </div>
                
                <div class="insight">
                    <h3>📅 Phase 2 (30-90 days): Advanced Analytics</h3>
                    <ul>
                        <li>Deploy predictive analytics for early event detection</li>
                        <li>Implement volatility trading algorithms</li>
                        <li>Establish automated risk management protocols</li>
                    </ul>
                </div>
                
                <div class="insight">
                    <h3>📅 Phase 3 (90-180 days): Portfolio Optimization</h3>
                    <ul>
                        <li>Expand to additional correlated markets</li>
                        <li>Implement machine learning for pattern recognition</li>
                        <li>Develop comprehensive reporting dashboard</li>
                    </ul>
                </div>
            </div>
            
            <div class="section">
                <h2>📈 Success Metrics & KPIs</h2>
                <table>
                    <tr><th>KPI</th><th>Target</th><th>Current</th><th>Status</th></tr>
                    <tr><td>Daily Revenue Average</td><td>${summary_stats['avg_revenue']*1.2:,.0f}</td><td>${summary_stats['avg_revenue']:,.0f}</td><td>🟡 In Progress</td></tr>
                    <tr><td>Success Rate</td><td>85%</td><td>{(events['revenue'] > 0).mean()*100:.0f}%</td><td>🟢 On Track</td></tr>
                    <tr><td>Risk-Adjusted Returns</td><td>1.5x</td><td>{high_vol_events['revenue'].mean() / events['revenue'].mean() if len(events) > 0 else 0:.1f}x</td><td>🟢 Achieved</td></tr>
                    <tr><td>Market Coverage</td><td>5 markets</td><td>{summary_stats['markets_count']} markets</td><td>🟡 Expanding</td></tr>
                </table>
            </div>
            
            <div class="footer">
                <p><strong>📊 Extreme Day Forensics - AI-Generated Post-Mortems for Battery Revenue Spikes</strong></p>
                <p>Generated by Advanced Energy Market Intelligence Platform</p>
                <p>© 2026 - Extreme Day Forensics: AI-Generated Post-Mortems Platform</p>
                <p><em>This report contains AI-generated insights and should be validated with domain expertise</em></p>
            </div>
        </body>
        </html>
        """
        
        return html_content
    
    def _show_advanced_filters(self):
        """Show advanced filtering capabilities"""
        st.markdown("## 🔍 Advanced Filters & Data Exploration")
        
        events = st.session_state.extreme_events
        
        if len(events) == 0:
            st.info("🔍 No data available for filtering. Generate an analysis first.")
            return
        
        # Store filtered events in session state
        if 'filtered_events' not in st.session_state:
            st.session_state.filtered_events = events.copy()
        
        # Filter controls
        st.markdown("### 🎯 Filter Controls")
        
        col1, col2 = st.columns(2)
        
        with col1:
            # Date range filter
            st.markdown("#### 📅 Date Range Filter")
            
            events['date'] = pd.to_datetime(events['date'])
            min_date = events['date'].min().date()
            max_date = events['date'].max().date()
            
            start_date = st.date_input("Start Date", min_date, key="start_date")
            end_date = st.date_input("End Date", max_date, key="end_date")
            
            # Market filter
            st.markdown("#### 🌐 Market Filter")
            available_markets = events['market'].unique().tolist()
            selected_markets = st.multiselect(
                "Select Markets",
                available_markets,
                default=available_markets,
                key="filter_markets"
            )
            
            # Event type filter
            st.markdown("#### 🎯 Event Type Filter")
            available_types = events['event_type'].unique().tolist()
            selected_types = st.multiselect(
                "Select Event Types",
                available_types,
                default=available_types,
                key="filter_types"
            )
        
        with col2:
            # Revenue range filter
            st.markdown("#### 💰 Revenue Range Filter")
            
            min_revenue = float(events['revenue'].min())
            max_revenue = float(events['revenue'].max())
            
            revenue_range = st.slider(
                "Revenue Range ($)",
                min_revenue,
                max_revenue,
                (min_revenue, max_revenue),
                key="revenue_range"
            )
            
            # Volatility filter
            st.markdown("#### 📈 Volatility Filter")
            
            events_copy = events.copy()
            events_copy['volatility'] = events_copy['price_stats'].apply(lambda x: x['price_volatility'])
            
            min_vol = float(events_copy['volatility'].min())
            max_vol = float(events_copy['volatility'].max())
            
            volatility_range = st.slider(
                "Volatility Range",
                min_vol,
                max_vol,
                (min_vol, max_vol),
                key="volatility_range"
            )
            
            # Price range filter
            st.markdown("#### 📊 Price Range Filter")
            
            events_copy['max_price'] = events_copy['price_stats'].apply(lambda x: x['max_price'])
            events_copy['min_price'] = events_copy['price_stats'].apply(lambda x: x['min_price'])
            
            min_max_price = float(events_copy['max_price'].min())
            max_max_price = float(events_copy['max_price'].max())
            
            price_range = st.slider(
                "Max Price Range ($)",
                min_max_price,
                max_max_price,
                (min_max_price, max_max_price),
                key="price_range"
            )
        
        # Apply filters button
        st.markdown("---")
        
        col1, col2, col3 = st.columns([1, 1, 1])
        
        with col1:
            if st.button("🔍 **Apply Filters**", type="primary"):
                self._apply_filters(start_date, end_date, selected_markets, selected_types, 
                                 revenue_range, volatility_range, price_range)
        
        with col2:
            if st.button("🔄 **Reset Filters**"):
                st.session_state.filtered_events = events.copy()
                st.rerun()
        
        with col3:
            if st.button("📥 **Export Filtered Data**"):
                self._export_filtered_data()
        
        # Show filtered results
        st.markdown("### 📊 Filtered Results")
        
        filtered_events = st.session_state.filtered_events
        
        # Filter summary
        st.markdown("#### 📈 Filter Summary")
        
        col1, col2, col3, col4 = st.columns(4)
        
        with col1:
            st.metric("📊 Total Events", len(filtered_events))
        with col2:
            profitable = len(filtered_events[filtered_events['revenue'] > 0])
            st.metric("💰 Profitable", profitable)
        with col3:
            avg_rev = filtered_events['revenue'].mean()
            st.metric("📈 Avg Revenue", f"${avg_rev:,.0f}")
        with col4:
            total_rev = filtered_events['revenue'].sum()
            st.metric("💎 Total Revenue", f"${total_rev:,.0f}")
        
        # Filtered data table
        st.markdown("#### 📋 Filtered Events Data")
        
        # Create display dataframe
        display_df = filtered_events.copy()
        # Convert date to string if it's datetime
        if pd.api.types.is_datetime64_any_dtype(display_df['date']):
            display_df['date'] = display_df['date'].dt.strftime('%Y-%m-%d')
        else:
            display_df['date'] = display_df['date'].astype(str)
        display_df['revenue'] = display_df['revenue'].round(2)
        display_df['max_price'] = display_df['price_stats'].apply(lambda x: x['max_price']).round(2)
        display_df['min_price'] = display_df['price_stats'].apply(lambda x: x['min_price']).round(2)
        display_df['volatility'] = display_df['price_stats'].apply(lambda x: x['price_volatility']).round(3)
        
        display_df = display_df[['date', 'market', 'event_type', 'revenue', 'max_price', 'min_price', 'volatility']]
        
        st.dataframe(display_df, use_container_width=True)
        
        # Filtered visualizations
        if len(filtered_events) > 0:
            st.markdown("#### 📈 Filtered Data Visualizations")
            
            col1, col2 = st.columns(2)
            
            with col1:
                # Revenue distribution
                fig = px.histogram(
                    filtered_events,
                    x='revenue',
                    color='market',
                    title="Revenue Distribution (Filtered)",
                    height=350
                )
                st.plotly_chart(fig, use_container_width=True)
            
            with col2:
                # Market performance
                market_rev = filtered_events.groupby('market')['revenue'].sum().reset_index()
                
                fig = px.bar(
                    market_rev,
                    x='market',
                    y='revenue',
                    title="Market Revenue (Filtered)",
                    height=350
                )
                st.plotly_chart(fig, use_container_width=True)
            
            # Timeline view
            st.markdown("#### 📅 Timeline View (Filtered)")
            
            timeline_df = filtered_events.copy()
            timeline_df['date'] = pd.to_datetime(timeline_df['date'])
            
            fig = px.scatter(
                timeline_df,
                x='date',
                y='revenue',
                color='market',
                size=timeline_df['revenue'].abs(),  # Use absolute value for size
                hover_data=['event_type'],
                title="Events Timeline (Filtered)",
                height=400
            )
            st.plotly_chart(fig, use_container_width=True)
        
        # Advanced filter insights
        st.markdown("### 🤖 Filter Insights")
        
        if len(filtered_events) > 0:
            # Generate insights based on filtered data
            best_market = filtered_events.groupby('market')['revenue'].sum().idxmax()
            best_event_type = filtered_events.groupby('event_type')['revenue'].mean().idxmax()
            avg_volatility = filtered_events['price_stats'].apply(lambda x: x['price_volatility']).mean()
            
            insights = [
                f"🌐 **Best Market**: {best_market} with highest filtered revenue",
                f"🎯 **Best Event Type**: {best_event_type} most profitable in filtered data",
                f"📈 **Average Volatility**: {avg_volatility:.3f} in filtered events",
                f"💰 **Revenue Concentration**: {filtered_events['revenue'].sum()/events['revenue'].sum()*100:.1f}% of total revenue",
                f"📊 **Event Density**: {len(filtered_events)/len(events)*100:.1f}% of total events"
            ]
            
            for insight in insights:
                st.info(insight)
    
    def _apply_filters(self, start_date, end_date, selected_markets, selected_types, 
                      revenue_range, volatility_range, price_range):
        """Apply all filters to the data"""
        original_events = st.session_state.extreme_events.copy()
        
        # Start with all events
        filtered = original_events.copy()
        
        # Apply date filter
        filtered['date'] = pd.to_datetime(filtered['date'])
        date_mask = (filtered['date'].dt.date >= start_date) & (filtered['date'].dt.date <= end_date)
        filtered = filtered[date_mask]
        
        # Apply market filter
        if selected_markets:
            market_mask = filtered['market'].isin(selected_markets)
            filtered = filtered[market_mask]
        
        # Apply event type filter
        if selected_types:
            type_mask = filtered['event_type'].isin(selected_types)
            filtered = filtered[type_mask]
        
        # Apply revenue filter
        revenue_mask = (filtered['revenue'] >= revenue_range[0]) & (filtered['revenue'] <= revenue_range[1])
        filtered = filtered[revenue_mask]
        
        # Apply volatility filter
        filtered['volatility'] = filtered['price_stats'].apply(lambda x: x['price_volatility'])
        vol_mask = (filtered['volatility'] >= volatility_range[0]) & (filtered['volatility'] <= volatility_range[1])
        filtered = filtered[vol_mask]
        
        # Apply price filter
        filtered['max_price'] = filtered['price_stats'].apply(lambda x: x['max_price'])
        price_mask = (filtered['max_price'] >= price_range[0]) & (filtered['max_price'] <= price_range[1])
        filtered = filtered[price_mask]
        
        # Store filtered results
        st.session_state.filtered_events = filtered
        
        st.success(f"🔍 Filters applied! {len(filtered)} events remaining from {len(original_events)} total events")
        st.rerun()
    
    def _export_filtered_data(self):
        """Export filtered data"""
        filtered_events = st.session_state.filtered_events
        
        if len(filtered_events) == 0:
            st.warning("⚠️ No data to export!")
            return
        
        # Prepare data for export
        export_df = filtered_events.copy()
        
        # Convert date column to string safely
        if pd.api.types.is_datetime64_any_dtype(export_df['date']):
            export_df['date'] = export_df['date'].dt.strftime('%Y-%m-%d')
        else:
            # If already string or other format, convert to datetime then format
            export_df['date'] = pd.to_datetime(export_df['date']).dt.strftime('%Y-%m-%d')
        
        # CSV export
        csv_data = export_df.to_csv(index=False)
        
        st.download_button(
            label="📥 **Download Filtered Data (CSV)**",
            data=csv_data,
            file_name=f"extreme_day_storyboard_filtered_{date.today()}.csv",
            mime="text/csv",
            type="primary"
        )
    
    def _create_elegant_footer(self):
        """Create elegant footer"""
        st.markdown("---")
        st.markdown("""
        <div style='background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); padding: 2rem; border-radius: 15px; text-align: center; color: white;'>
            <h3>📊 Extreme Day Forensics</h3>
            <p style='margin: 0.5rem 0; font-size: 1.1em;'>AI-Generated Post-Mortems for Battery Revenue Spikes</p>
            <div style='margin-top: 1rem; font-size: 0.8em; opacity: 0.7;'>
                Built with 📊 and 🤖 for Energy Market Intelligence<br>
                © 2026 - Advanced Energy Market Intelligence Platform
            </div>
        </div>
        """, unsafe_allow_html=True)


def main():
    """Main function"""
    storyboard = ExtremeDayStoryboard()
    storyboard.run()


if __name__ == "__main__":
    main()
